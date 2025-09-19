#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
从 Word 文档中提取书籍信息并写入 Excel
字段顺序：编号、书名、出版社、作者、ISBN、定价、实洋
"""

from docx import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import os
import sys
from typing import List, Dict

# ----------------- 配置 -----------------
WORD_FILE = "D:\python_about\demo\订书\附录1.docx"       # 输入文件
EXCEL_FILE = "D:\python_about\demo\订书\附录2.xlsx"     # 输出文件
# ---------------------------------------

# 需要提取的字段及其前缀
FIELD_PREFIX = {
    "书名": "书名",
    "出版社": "出版社",
    "作者": "作者",
    "ISBN": "ISBN",
    "定价": "定价",
    "实洋": "实洋"
}

def parse_word(file_path: str) -> List[Dict[str, str]]:
    """
    解析 Word 文档，返回书籍列表
    每个元素为 dict，键为字段名
    """
    if not os.path.exists(file_path):
        sys.exit(f"文件不存在：{file_path}")

    doc = Document(file_path)
    books = []
    current = None

    for para in doc.paragraphs:
        line = para.text.strip()
        if not line:
            continue

        # 判断是否为新条目的开始
        if line.startswith("书名："):
            if current:  # 保存上一条
                books.append(current)
            current = {"书名": "", "出版社": "", "作者": "", "ISBN": "", "定价": "", "实洋": ""}

        # 尝试匹配 6 个字段
        for field, prefix in FIELD_PREFIX.items():
            if line.startswith(f"{prefix}："):
                value = line.split("：", 1)[1].strip()
                if current is None:
                    # 容错：如果文档开头没有“书名”就开始出现其它字段，则自动创建空字典
                    current = {"书名": "", "出版社": "", "作者": "", "ISBN": "", "定价": "", "实洋": ""}
                current[field] = value
                break

    if current:  # 最后一条
        books.append(current)

    return books

def write_excel(books: List[Dict[str, str]], excel_path: str):
    """
    将书籍信息写入 Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "书籍信息"

    # 表头
    headers = ["编号", "书名", "出版社", "作者", "ISBN", "定价", "实洋"]
    ws.append(headers)

    # 写入数据
    for idx, book in enumerate(books, start=1):
        row = [
            idx,
            book["书名"],
            book["出版社"],
            book["作者"],
            book["ISBN"],
            book["定价"],
            book["实洋"]
        ]
        ws.append(row)

    # 简单美化：自动列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # 列号（1 开始）
        col_letter = get_column_letter(column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"已生成 Excel：{excel_path}")

def main():
    books = parse_word(WORD_FILE)
    if not books:
        print("未提取到任何书籍信息，请检查 Word 文件格式。")
        return
    write_excel(books, EXCEL_FILE)

if __name__ == "__main__":
    main()